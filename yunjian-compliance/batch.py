# -*- coding: utf-8 -*-
"""
批量言论分析模块
==============
解析 CSV / Excel 言论表（每行一条言论，可含姓名/网点），逐条分析
（情感 + 合规），并聚合为统计看板所需的结构化数据。

表格式：
  言论 | 姓名 | 网点            （列名支持中文/英文，列可缺省）
  ……    ……     ……

聚合输出见 aggregate() 的返回结构说明。
"""
import csv
import io

import compliance
import sentiment

# 支持的表头（大小写不敏感）
TEXT_COLS = {"言论", "内容", "文本", "话术", "意见", "反馈", "评语",
             "评论", "留言", "说话内容", "text", "content", "remark",
             "feedback", "comment", "speech"}
GROUP_COLS = {"网点", "部门", "支行", "机构", "营业部", "分支", "组",
              "branch", "dept", "org", "group"}
NAME_COLS = {"姓名", "员工", "员工姓名", "名字", "工号", "name", "employee", "user"}


def read_table(path):
    """读取 CSV / XLSX，返回行列表 [[...], ...]（含表头行则原样返回）。"""
    lower = path.lower()
    if lower.endswith((".xlsx", ".xlsm")):
        return _read_xlsx(path)
    if lower.endswith(".xls"):
        raise ValueError("暂不支持旧版 .xls，请另存为 .xlsx 或 .csv")
    return _read_csv(path)


def _read_csv(path):
    raw = None
    for enc in ("utf-8-sig", "gbk", "gb18030"):
        try:
            with io.open(path, "r", encoding=enc, newline="") as fh:
                raw = fh.read()
            break
        except (UnicodeDecodeError, OSError):
            continue
    if raw is None:
        with io.open(path, "r", encoding="utf-8", errors="ignore", newline="") as fh:
            raw = fh.read()
    rows = list(csv.reader(io.StringIO(raw)))
    return [[str(c).strip() for c in row] for row in rows if any(c.strip() for c in row)]


def _read_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(values_only=True):
        cells = ["" if c is None else str(c).strip() for c in row]
        if any(cells):
            rows.append(cells)
    wb.close()
    return rows


def _find_header_index(headers, cols):
    for i, h in enumerate(headers):
        if h.strip().lower() in cols:
            return i
    return -1


def parse_rows(table):
    """把读到的表格解析为 [{name, group, text}, ...]。

    表头识别失败时，按「首列=言论、次列=网点」兜底。
    """
    if not table:
        return []

    first = table[0]
    text_i = _find_header_index(first, TEXT_COLS)
    group_i = _find_header_index(first, GROUP_COLS)
    name_i = _find_header_index(first, NAME_COLS)
    has_header = (text_i >= 0 or group_i >= 0 or name_i >= 0)

    data_rows = table[1:] if has_header else table
    if not has_header:
        # 无表头：首列言论、次列网点
        text_i, group_i, name_i = 0, 1, -1

    items = []
    for row in data_rows:
        text = row[text_i].strip() if text_i < len(row) else ""
        if not text:
            continue
        group = row[group_i].strip() if (group_i >= 0 and group_i < len(row)) else ""
        name = row[name_i].strip() if (name_i >= 0 and name_i < len(row)) else ""
        items.append({"name": name, "group": group, "text": text})
    return items


def analyze_batch(items):
    """逐条分析，返回带结果的条目列表。"""
    out = []
    for it in items:
        res = sentiment.analyze(it["text"])
        hits = compliance.detect(it["text"])
        has_high = any(h["severity"] == "高" for h in hits)
        out.append({
            "name": it["name"],
            "group": it["group"],
            "text": it["text"],
            "label": res["label"],
            "risk": res["negative_risk"],
            "categories": [h["category"] for h in hits],
            "severities": [h["severity"] for h in hits],
            "compliance": hits,
            "has_high": has_high,
        })
    return out


def aggregate(items):
    """聚合为统计看板数据。

    返回 dict：
      total           总条数
      negative_count  负面条数
      compliance_rows 命中合规风险的条数
      high_rows       命中高风险（等级=高）的条数
      sentiment       情感分布 {标签: 条数}
      severity        风险等级命中分布 {高/中/低: 条次}
      categories      合规类别统计（含建议，供明细表）
      branches        网点统计（供排名图）
      high_risk_items 高风险言论明细（供明细表）
      has_group       是否含网点/分组信息
    """
    total = len(items)
    negative_count = sum(1 for x in items if x["label"] == "负面")

    sentiment_dist = {"正面": 0, "中性": 0, "负面": 0}
    for x in items:
        sentiment_dist[x["label"]] = sentiment_dist.get(x["label"], 0) + 1

    severity_dist = {"高": 0, "中": 0, "低": 0}
    cat_map = {}   # category -> {count, severity, suggestion}
    for x in items:
        for h in x["compliance"]:
            severity_dist[h["severity"]] = severity_dist.get(h["severity"], 0) + 1
            c = cat_map.setdefault(h["category"], {
                "category": h["category"],
                "count": 0,
                "severity": h["severity"],
                "suggestion": h["suggestion"],
            })
            c["count"] += 1

    compliance_rows = sum(1 for x in items if x["categories"])
    high_rows = sum(1 for x in items if x["has_high"])
    has_name = any(x["name"] for x in items)

    # 网点统计
    has_group = any(x["group"] for x in items)
    branches = {}
    for x in items:
        g = x["group"] or "（未标注）"
        b = branches.setdefault(g, {"group": g, "total": 0, "negative": 0,
                                    "compliance": 0, "high": 0})
        b["total"] += 1
        b["negative"] += (1 if x["label"] == "负面" else 0)
        b["compliance"] += (1 if x["categories"] else 0)
        b["high"] += (1 if x["has_high"] else 0)
    branch_list = sorted(branches.values(), key=lambda b: (-b["high"], -b["compliance"]))

    # 高风险言论明细（按风险分降序，最多取 30 条）
    risky = [x for x in items if x["categories"]]
    risky.sort(key=lambda x: (-x["risk"], -len(x["categories"])))
    high_risk_items = [{
        "idx": i + 1,
        "name": x["name"] or "—",
        "group": x["group"] or "—",
        "text": x["text"][:80],
        "categories": x["categories"],
        "risk": x["risk"],
        "label": x["label"],
    } for i, x in enumerate(risky[:30])]

    return {
        "total": total,
        "negative_count": negative_count,
        "compliance_rows": compliance_rows,
        "high_rows": high_rows,
        "sentiment": sentiment_dist,
        "severity": severity_dist,
        "categories": sorted(cat_map.values(), key=lambda c: -c["count"]),
        "branches": branch_list,
        "has_group": has_group,
        "has_name": has_name,
        "high_risk_items": high_risk_items,
    }
